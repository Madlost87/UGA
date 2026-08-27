from __future__ import annotations

import re
import shutil
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.utils import ImageReader
from reportlab.graphics.shapes import Drawing, Polygon
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Flowable, Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from PIL import Image as PILImage

from config import (
    BOTANICAL_WATERCOLOR,
    DISPLAY_HEADERS,
    DEFAULT_INPUT_FILE,
    DOWNLOADED_PAPER_TEXTURE,
    FONTS_DIR,
    GENERATED_PAPER_TEXTURE,
    ICONS_DIR,
    INPUT_FILE,
    OUTPUT_FILE,
    PAPER_TEXTURE,
    PREVIEW_FILE,
    REQUIRED_COLUMNS,
    SHEET_NAME,
)
from style import STYLE
from artwork import ensure_legend_icons, generate_paper_artwork


class MenuGenerationError(Exception):
    pass


@dataclass(frozen=True)
class RecipeRow:
    values: dict[str, str]


@dataclass(frozen=True)
class LegendEntry:
    symbol_type: str
    label: str
    description: str


class WatermarkedParagraph(Flowable):
    def __init__(self, paragraph: Any, image_path: Path, size: float, opacity: float) -> None:
        super().__init__()
        self.paragraph = paragraph
        self.image_path = image_path
        image = PILImage.open(image_path).convert("RGBA")
        bbox = image.getchannel("A").getbbox()
        if bbox:
            image = image.crop(bbox)
        self.image = ImageReader(image)
        self.image_width = image.width
        self.image_height = image.height
        self.size = size
        self.opacity = opacity
        self.width = 0
        self.height = 0

    def draw_dimensions(self, max_width: float, max_height: float) -> tuple[float, float]:
        scale = min(self.size / max(self.image_width, self.image_height), max_width / self.image_width, max_height / self.image_height)
        return self.image_width * scale, self.image_height * scale

    def wrap(self, availWidth: float, availHeight: float) -> tuple[float, float]:
        _paragraph_width, paragraph_height = self.paragraph.wrap(availWidth, availHeight)
        self.width = availWidth
        self.height = paragraph_height
        return availWidth, paragraph_height

    def draw(self) -> None:
        self.canv.saveState()
        if hasattr(self.canv, "setFillAlpha"):
            self.canv.setFillAlpha(self.opacity)
        inset = STYLE["watermark"]["corner_inset"]
        draw_width, draw_height = self.draw_dimensions(max(self.width - inset * 2, 1), max(self.height - inset * 2, 1))
        x = max(self.width - draw_width - inset, 0)
        y = inset
        self.canv.drawImage(self.image, x, y, width=draw_width, height=draw_height, mask="auto")
        self.canv.restoreState()
        self.paragraph.drawOn(self.canv, 0, 0)


class RecipeTable(Table):
    def _drawCell(self, cellval: Any, cellstyle: Any, pos: tuple[float, float], size: tuple[float, float]) -> None:
        if isinstance(cellval, WatermarkedParagraph):
            colpos, rowpos = pos
            colwidth, rowheight = size
            inset = STYLE["watermark"]["corner_inset"]
            draw_width, draw_height = cellval.draw_dimensions(max(colwidth - inset * 2, 1), max(rowheight - inset * 2, 1))
            x = colpos + colwidth - draw_width - inset
            y = rowpos + inset
            self.canv.saveState()
            if hasattr(self.canv, "setFillAlpha"):
                self.canv.setFillAlpha(cellval.opacity)
            self.canv.drawImage(cellval.image, x, y, width=draw_width, height=draw_height, mask="auto")
            self.canv.restoreState()
            return super()._drawCell(cellval.paragraph, cellstyle, pos, size)
        return super()._drawCell(cellval, cellstyle, pos, size)


def hex_color(value: str) -> colors.Color:
    return colors.HexColor(value)


def register_fonts() -> None:
    fonts = STYLE["fonts"]
    standard_fonts = {
        "Courier",
        "Courier-Bold",
        "Courier-Oblique",
        "Courier-BoldOblique",
        "Helvetica",
        "Helvetica-Bold",
        "Helvetica-Oblique",
        "Helvetica-BoldOblique",
        "Times-Roman",
        "Times-Bold",
        "Times-Italic",
        "Times-BoldItalic",
        "Symbol",
        "ZapfDingbats",
    }
    candidates = {
        fonts["regular"]: [
            str(FONTS_DIR / "Kalam-Regular.ttf"),
            "/usr/share/fonts/noto/NotoSerif-Regular.ttf",
            "/usr/share/fonts/TTF/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/local/share/fonts/DejaVuSans.ttf",
        ],
        fonts["bold"]: [
            str(FONTS_DIR / "Kalam-Bold.ttf"),
            "/usr/share/fonts/noto/NotoSerif-Bold.ttf",
            "/usr/share/fonts/TTF/DejaVuSans-Bold.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
            "/usr/local/share/fonts/DejaVuSans-Bold.ttf",
        ],
        fonts["title"]: [
            str(FONTS_DIR / "CaveatBrush-Regular.ttf"),
            "/usr/share/fonts/noto/NotoSerifDisplay-SemiBoldItalic.ttf",
            "/usr/share/fonts/TTF/DejaVuSerif-Bold.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSerif-Bold.ttf",
            "/usr/share/fonts/liberation/LiberationSerif-Bold.ttf",
        ],
        fonts["hand"]: [
            str(FONTS_DIR / "CaveatBrush-Regular.ttf"),
            "/usr/share/fonts/noto/NotoSerifDisplay-Italic.ttf",
            "/usr/share/fonts/TTF/DejaVuSerif-Italic.ttf",
            "/usr/share/fonts/liberation/LiberationSerif-Italic.ttf",
        ],
        fonts["symbols"]: [
            "/usr/share/fonts/TTF/DejaVuSans.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
            "/usr/local/share/fonts/DejaVuSans.ttf",
        ],
    }

    for font_name, paths in candidates.items():
        if font_name in standard_fonts:
            continue
        if font_name in pdfmetrics.getRegisteredFontNames():
            continue
        for path in paths:
            if Path(path).exists():
                pdfmetrics.registerFont(TTFont(font_name, path))
                break


def font_name(kind: str) -> str:
    fonts = STYLE["fonts"]
    requested = fonts[kind]
    standard_fonts = {
        "Courier",
        "Courier-Bold",
        "Courier-Oblique",
        "Courier-BoldOblique",
        "Helvetica",
        "Helvetica-Bold",
        "Helvetica-Oblique",
        "Helvetica-BoldOblique",
        "Times-Roman",
        "Times-Bold",
        "Times-Italic",
        "Times-BoldItalic",
        "Symbol",
        "ZapfDingbats",
    }
    if requested in standard_fonts:
        return requested
    if requested in pdfmetrics.getRegisteredFontNames():
        return requested
    if kind == "title":
        return fonts["fallback_title"]
    if kind == "hand":
        return fonts["fallback_hand"]
    return fonts["fallback_bold"] if kind == "bold" else fonts["fallback_regular"]


def normalize_header(value: Any) -> str:
    return str(value or "").strip().upper()


def clean_cell_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return normalize_text(str(value))


def normalize_text(value: str) -> str:
    replacements = {
        "\r\n": "\n",
        "\r": "\n",
        "→": "->",
        " piu ": " più ",
        " piu' ": " più ",
        "Piu ": "Più ",
        "mov.": "movimento",
        "mov ": "movimento ",
    }
    text = value.strip()
    ability_pattern = r"(?<!\w)(?:(\d+)\s*)?((?:U|◎)+)\s*(↪|⚔|O|▽|✱)\s*:?"
    text = re.sub(ability_pattern, normalize_ability_marker, text)
    text = re.sub(r"(?:●[ \t]*\d+[ \t]*)+", format_ring_markers, text)
    text = re.sub(r"\bG\s*(\d+)", r"G Gittata: \1", text)
    text = re.sub(r"(\d)([A-Za-zÀ-ÿ])", r"\1 \2", text)
    text = re.sub(r"([a-zà-ÿ])([A-ZÀ-Ý])", r"\1 \2", text)
    text = re.sub(r"●+", "Azione", text)
    text = text.replace("◎", "U")
    for source, target in replacements.items():
        text = text.replace(source, target)

    lines = []
    for raw_line in text.split("\n"):
        line = re.sub(r"\s+", " ", raw_line).strip()
        if not line:
            continue
        if line.upper() == "X":
            line = STYLE["text"]["empty_cell"]
        if STYLE["text"]["strip_unreadable_notes"] and line.lower().startswith("[nota barrata"):
            continue
        line = line.replace("Accessorio: NO", "Accessorio: no")
        line = line.replace("accessorio x", "accessorio: ")
        line = line.replace("Accessorio x", "Accessorio: ")
        line = line.replace("accessorio >", "accessorio: ")
        line = line.replace("Accessorio >", "Accessorio: ")
        line = line.replace("Movimento movimento", "Movimento")
        line = line.replace("movimento movimento", "movimento")
        line = re.sub(r"\b(Movimento|Attacco|Difesa|Estrazione|Speciale):\s*\1\b", r"\1:", line, flags=re.IGNORECASE)
        line = re.sub(r"\b(Movimento|Attacco|Difesa|Estrazione|Speciale|Accessorio|Termina|Reazione|Gittata):\s*", r"\1: ", line, flags=re.IGNORECASE)
        line = re.sub(r"\b(\d+)\s+U\s+([↪⚔O▽✱])", r"\1U \2", line)
        line = re.sub(r"^-\s*(accendi|spegni)\s*\((.*?)\)$", lambda m: f"{m.group(1).capitalize()}: {m.group(2)}", line, flags=re.IGNORECASE)
        line = re.sub(r"^-\s*(accendi|spegni)\s*$", lambda m: f"{m.group(1).capitalize()}", line, flags=re.IGNORECASE)
        line = line.replace("AzioneAzione", "Azione")
        line = line.replace("Usospeciale", "Uso speciale")
        line = line.replace("danno allontana?", "danno: allontana?")
        line = line.replace("Termina: movimento non azione", "Termina: movimento, non azione")
        line = line.replace("linee retta", "linea retta")
        line = re.sub(r"\bin R\b", "in Raggio", line)
        line = line.replace("un azione", "un'azione")
        line = line.replace("in Anello", "a raggio")
        lines.append(prettify_line(line))
    return "\n".join(lines)


def normalize_ability_marker(match: re.Match[str]) -> str:
    explicit_count = match.group(1)
    use_token = match.group(2)
    symbol = match.group(3)
    if explicit_count:
        count = int(explicit_count)
    elif "◎" in use_token:
        count = use_token.count("◎")
    else:
        count = max(1, use_token.upper().count("U"))

    labels = {
        "↪": "Movimento",
        "⚔": "Attacco",
        "O": "Estrazione",
        "▽": "Abilità difesa",
        "✱": "Speciale",
    }
    uses = "U" if count == 1 else f"{count}U"
    return f"{uses} {symbol} {labels[symbol]}:"


def format_ring_markers(match: re.Match[str]) -> str:
    numbers = re.findall(r"\d+", match.group(0))
    if not numbers:
        return ""
    label = "Raggio" if len(numbers) == 1 else "Raggi"
    return f"{label} {'/'.join(numbers)}"


def prettify_line(line: str) -> str:
    word_replacements = {
        "ACCESSORIO": "Accessorio",
        "ACCESSORI": "Accessori",
        "AZIONI": "azioni",
        "ACCENDI": "Accendi",
        "SPEGNI": "Spegni",
        "PIETRA": "pietra",
        "PELLE": "pelle",
        "RAMO": "ramo",
        "BACCA": "bacca",
        "BASTONE": "bastone",
        "LASTRA": "lastra",
    }
    pretty = line.replace("->", " -> ")
    pretty = re.sub(r"\s+", " ", pretty).strip()
    for source, target in word_replacements.items():
        pretty = re.sub(rf"\b{source}\b", target, pretty)
    if pretty and pretty[0].islower():
        pretty = pretty[0].upper() + pretty[1:]
    return pretty


def validate_columns(headers: list[Any]) -> dict[str, int | None]:
    mapping = {normalize_header(header): index for index, header in enumerate(headers)}
    optional = {"U_POTENZIAMENTO_3"}
    missing = [column for column in REQUIRED_COLUMNS if column not in mapping and column not in optional]
    if missing:
        missing_text = ", ".join(missing)
        raise MenuGenerationError(f"Colonne mancanti nel file Excel: {missing_text}")
    return {column: mapping.get(column) for column in REQUIRED_COLUMNS}


def find_header_index(headers: list[str], aliases: set[str], start: int = 0) -> int | None:
    for index in range(start, len(headers)):
        if headers[index] in aliases:
            return index
    return None


def validate_uploaded_a3_columns(headers: list[Any]) -> dict[str, int | None]:
    normalized = [normalize_header(header) for header in headers]

    creation = find_header_index(normalized, {"CREAZIONI", "CREAZIONE"})
    materials = find_header_index(normalized, {"MATERIALI"})
    upgrade_1 = find_header_index(normalized, {"POTENZIAMENTO 1", "POTENZIAMENTO_1", "POTENZIAMENTO I"})
    upgrade_2 = find_header_index(normalized, {"POTENZIAMENTO 2", "POTENZIAMENTO_2", "POTENZIAMENTO II"})
    upgrade_3 = find_header_index(normalized, {"POTENZIAMENTO 3", "POTENZIAMENTO_3", "POTENZIAMENTO III"})
    notes = find_header_index(normalized, {"AZIONE SPECIALE / NOTE", "NOTE", "AZIONE SPECIALE"})

    required = {
        "CREAZIONI": creation,
        "MATERIALI": materials,
        "POTENZIAMENTO 1": upgrade_1,
        "POTENZIAMENTO 2": upgrade_2,
        "POTENZIAMENTO 3": upgrade_3,
    }
    missing = [name for name, index in required.items() if index is None]
    if missing:
        missing_text = ", ".join(missing)
        raise MenuGenerationError(f"Colonne mancanti nel file Excel: {missing_text}")

    u_after_creation = find_header_index(normalized, {"U"}, creation + 1 if creation is not None else 0)
    u_after_upgrade_1 = find_header_index(normalized, {"U"}, upgrade_1 + 1 if upgrade_1 is not None else 0)
    u_after_upgrade_2 = find_header_index(normalized, {"U"}, upgrade_2 + 1 if upgrade_2 is not None else 0)
    u_after_upgrade_3 = find_header_index(normalized, {"U"}, upgrade_3 + 1 if upgrade_3 is not None else 0)

    return {
        "CREAZIONE": creation,
        "U_CREAZIONE": u_after_creation,
        "MATERIALI": materials,
        "U_MATERIALI": None,
        "POTENZIAMENTO_1": upgrade_1,
        "U_POTENZIAMENTO_1": u_after_upgrade_1,
        "POTENZIAMENTO_2": upgrade_2,
        "U_POTENZIAMENTO_2": u_after_upgrade_2,
        "POTENZIAMENTO_3": upgrade_3,
        "U_POTENZIAMENTO_3": u_after_upgrade_3,
        "_NOTE": notes,
    }


def resolve_column_map(headers: list[Any]) -> dict[str, int | None]:
    normalized = {normalize_header(header) for header in headers}
    if set(REQUIRED_COLUMNS).issubset(normalized):
        return validate_columns(headers)
    return validate_uploaded_a3_columns(headers)


def row_value(raw_row: tuple[Any, ...], index: int | None) -> str:
    if index is None or index >= len(raw_row):
        return ""
    return clean_cell_value(raw_row[index])


def normalize_recipe_values(raw_row: tuple[Any, ...], column_map: dict[str, int | None]) -> dict[str, str]:
    values = {column: row_value(raw_row, column_map[column]) for column in REQUIRED_COLUMNS}
    return values


def is_recipe_row(values: dict[str, str]) -> bool:
    return bool(values["CREAZIONE"] and (values["U_CREAZIONE"] or values["MATERIALI"]))


def load_workbook_data(path: Path = INPUT_FILE, sheet_name: str | None = SHEET_NAME) -> list[RecipeRow]:
    if not path.exists():
        raise MenuGenerationError(f"File Excel non trovato: {path}")

    try:
        workbook = load_workbook(path, data_only=True)
    except InvalidFileException as exc:
        raise MenuGenerationError(f"Formato Excel non valido: {path}") from exc
    except OSError as exc:
        raise MenuGenerationError(f"Impossibile leggere il file Excel: {path}") from exc

    if not workbook.sheetnames:
        raise MenuGenerationError("Il file Excel non contiene fogli.")

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise MenuGenerationError(f"Foglio non trovato: {sheet_name}")
        sheet = workbook[sheet_name]
    else:
        sheet = workbook[workbook.sheetnames[0]]

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise MenuGenerationError("Il foglio Excel e vuoto.")

    column_map = resolve_column_map(list(rows[0]))
    recipes: list[RecipeRow] = []

    for raw_row in rows[1:]:
        values = normalize_recipe_values(raw_row, column_map)
        if is_recipe_row(values):
            recipes.append(RecipeRow(values=values))

    if not recipes:
        raise MenuGenerationError("Nessuna ricetta trovata: tutte le righe dati sono vuote.")

    return recipes


def load_footer_legend(path: Path = INPUT_FILE, sheet_name: str | None = SHEET_NAME) -> str:
    if not path.exists():
        raise MenuGenerationError(f"File Excel non trovato: {path}")

    try:
        workbook = load_workbook(path, data_only=True)
    except InvalidFileException as exc:
        raise MenuGenerationError(f"Formato Excel non valido: {path}") from exc
    except OSError as exc:
        raise MenuGenerationError(f"Impossibile leggere il file Excel: {path}") from exc

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise MenuGenerationError(f"Foglio non trovato: {sheet_name}")
        sheet = workbook[sheet_name]
    else:
        sheet = workbook[workbook.sheetnames[0]]

    for row in sheet.iter_rows(values_only=True):
        for value in row:
            text = "" if value is None else str(value).strip()
            if text.upper().startswith("LEGENDA"):
                return text

    raise MenuGenerationError("Legenda non trovata nel file Excel: serve una cella che inizi con 'LEGENDA'.")


def create_paragraph_styles() -> dict[str, ParagraphStyle]:
    cell = STYLE["cell"]
    header = STYLE["header"]
    creation = STYLE["creation"]
    quantity = STYLE["quantity"]

    return {
        "header": ParagraphStyle(
            "MenuHeader",
            fontName=font_name("title"),
            fontSize=header["font_size"],
            leading=header["leading"],
            textColor=hex_color(header["text_color"]),
            alignment=TA_CENTER,
            spaceAfter=0,
        ),
        "cell": ParagraphStyle(
            "MenuCell",
            fontName=font_name("regular"),
            fontSize=cell["font_size"],
            leading=cell["leading"],
            textColor=hex_color(cell["text_color"]),
            alignment=TA_LEFT,
            spaceAfter=0,
        ),
        "materials": ParagraphStyle(
            "MenuMaterials",
            fontName=font_name("regular"),
            fontSize=cell["font_size"],
            leading=cell["leading"],
            textColor=hex_color(STYLE["colors"]["materials_text"]),
            alignment=TA_LEFT,
            spaceAfter=0,
        ),
        "upgrade_1": ParagraphStyle(
            "MenuUpgrade1",
            fontName=font_name("regular"),
            fontSize=cell["font_size"],
            leading=cell["leading"],
            textColor=hex_color(STYLE["colors"]["upgrade_1_text"]),
            alignment=TA_LEFT,
            spaceAfter=0,
        ),
        "upgrade_2": ParagraphStyle(
            "MenuUpgrade2",
            fontName=font_name("regular"),
            fontSize=cell["font_size"],
            leading=cell["leading"],
            textColor=hex_color(STYLE["colors"]["upgrade_2_text"]),
            alignment=TA_LEFT,
            spaceAfter=0,
        ),
        "upgrade_3": ParagraphStyle(
            "MenuUpgrade3",
            fontName=font_name("regular"),
            fontSize=cell["font_size"],
            leading=cell["leading"],
            textColor=hex_color(STYLE["colors"]["upgrade_3_text"]),
            alignment=TA_LEFT,
            spaceAfter=0,
        ),
        "creation": ParagraphStyle(
            "MenuCreation",
            fontName=font_name("bold") if creation["bold"] else font_name("regular"),
            fontSize=creation["font_size"],
            leading=creation["leading"],
            textColor=hex_color(creation["text_color"]),
            alignment=TA_LEFT,
            spaceAfter=0,
        ),
        "quantity": ParagraphStyle(
            "MenuQuantity",
            fontName=font_name("bold") if quantity["bold"] else font_name("regular"),
            fontSize=quantity["font_size"],
            leading=quantity["leading"],
            textColor=hex_color(quantity["text_color"]),
            alignment=TA_CENTER,
            spaceAfter=0,
        ),
        "footer_title": ParagraphStyle(
            "MenuFooterTitle",
            fontName=font_name("title"),
            fontSize=STYLE["footer"]["title_size"],
            leading=STYLE["footer"]["title_leading"],
            textColor=hex_color(STYLE["footer"]["text_color"]),
            alignment=TA_CENTER,
            spaceAfter=0,
        ),
        "footer": ParagraphStyle(
            "MenuFooter",
            fontName=font_name("regular"),
            fontSize=STYLE["footer"]["font_size"],
            leading=STYLE["footer"]["leading"],
            textColor=hex_color(STYLE["footer"]["text_color"]),
            alignment=TA_LEFT,
            spaceAfter=0,
        ),
        "footer_symbol": ParagraphStyle(
            "MenuFooterSymbol",
            fontName=font_name("bold"),
            fontSize=STYLE["footer"]["font_size"],
            leading=STYLE["footer"]["leading"],
            textColor=hex_color(STYLE["footer"]["text_color"]),
            alignment=TA_CENTER,
            spaceAfter=0,
        ),
    }


def escape_text(value: str) -> str:
    return (
        value.replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace("\n", "<br/>")
    )


SECTION_LABELS = ("Costo", "Prerequisito", "Effetto", "Regola", "Note")


def format_section_breaks(value: str) -> str:
    section_labels = tuple(f"{label} -" for label in SECTION_LABELS)
    lines = value.split("\n")
    formatted: list[str] = []
    previous_section_label: str | None = None

    for line in lines:
        section_label = next((label for label in section_labels if line.startswith(label)), None)
        if formatted and section_label and previous_section_label and section_label != previous_section_label:
            formatted.append("")
        formatted.append(line)
        previous_section_label = section_label

    return "\n".join(formatted)


def split_section_line(line: str) -> tuple[str, str] | None:
    match = re.match(rf"^({'|'.join(SECTION_LABELS)})\s+-\s*(.*)$", line)
    if not match:
        return None
    return match.group(1), match.group(2).strip()


def section_hanging_indent(label: str, label_font: str, label_size: float, gap: float = 4.0) -> float:
    return pdfmetrics.stringWidth(f"{label}:", label_font, label_size) + gap


def semantic_highlight(value: str, bold_font: str | None = None, iconize_words: bool = True) -> str:
    semantic = STYLE["semantic_highlights"]
    if not semantic["enabled"] or not value:
        return escape_text(value)

    label_terms = semantic.get("label_terms", [])
    label_pattern = re.compile(
        rf"(?<!\w)({'|'.join(re.escape(term) for term in label_terms)})\s*:",
        flags=re.IGNORECASE,
    ) if label_terms else None

    protected: list[str] = []

    symbol_pattern = re.compile(
        r"(?<!\w)(?:\d+\s*)?U\s*[↪⚔▽✱O]"
        r"|(?<!\w)(?:\d+\s*)?[↪⚔▽✱O](?=\s+(?:Attacco|Movimento|Abilità\s+difesa|Estrazione|Speciale):)"
        r"|(?<!\w)G(?=\s+Gittata:)",
        flags=re.IGNORECASE,
    )
    inline_icon_pattern = re.compile(r"(?<!\w)(L3|ES|E5)(?!\w)", flags=re.IGNORECASE)
    inline_icon_files = {
        "L3": "l3.png",
        "ES": "es.png",
        "E5": "es.png",
    }
    word_icon_pattern = re.compile(r"(?<!\w)(raggio|raggi|accessorio|accessori)(?!\w)", flags=re.IGNORECASE)
    word_icon_files = {
        "raggio": "new_symbols/09_hex_2_2_2.png",
        "raggi": "new_symbols/09_hex_2_2_2.png",
        "accessorio": "new_symbols/06_wrench.png",
        "accessori": "new_symbols/06_wrench.png",
    }
    action_icon_files = {
        "↪": "movement.png",
        "⚔": "attack.png",
        "▽": "defense.png",
        "✱": "special.png",
        "O": "draw.png",
        "G": "range.png",
    }

    def icon_token(filename: str, size: float) -> str | None:
        icon_path = ICONS_DIR / filename
        if not icon_path.exists():
            return None
        return f'<img src="{icon_path}" width="{size}" height="{size}" valign="middle"/>'

    def replace_symbol(match: re.Match[str]) -> str:
        raw = match.group(0)
        use_match = re.match(r"(?P<count>\d+\s*)?U\s*(?P<action>[↪⚔▽✱O])", raw)
        size = STYLE["cell"]["font_size"] + 1.2
        if use_match:
            count = re.sub(r"\s+", "", use_match.group("count") or "")
            use_icon = icon_token("use.png", size)
            action_icon = icon_token(action_icon_files[use_match.group("action")], size)
            if use_icon and action_icon:
                token = (
                    f'<font name="{font_name("symbols")}" size="{size}" color="{semantic["symbol_color"]}">'
                    f'{escape_text(count)}</font>'
                    f'{use_icon}'
                    f' {action_icon}'
                )
                protected.append(token)
                return f"\uE000{len(protected) - 1}\uE001"

        standalone_match = re.match(r"(?P<count>\d+\s*)?(?P<symbol>[↪⚔▽✱O])", raw)
        if standalone_match:
            count = re.sub(r"\s+", "", standalone_match.group("count") or "")
            icon = icon_token(action_icon_files[standalone_match.group("symbol")], size)
            if icon:
                token = (
                    f'<font name="{font_name("bold")}" size="{size}" color="{semantic["symbol_color"]}">'
                    f'{escape_text(count)}</font>{icon}'
                )
                protected.append(token)
                return f"\uE000{len(protected) - 1}\uE001"

        range_icon = icon_token(action_icon_files["G"], size)
        if range_icon:
            protected.append(range_icon)
            return f"\uE000{len(protected) - 1}\uE001"

        return escape_text(raw)

    def replace_inline_icon(match: re.Match[str]) -> str:
        return ""

    def replace_word_icon(match: re.Match[str]) -> str:
        word = match.group(1).lower()
        if word in {"raggio", "raggi"}:
            return ""
        size = STYLE["cell"]["font_size"] * 3 if word in {"raggio", "raggi"} else STYLE["cell"]["font_size"] + 4.6
        icon = icon_token(word_icon_files[word], size)
        if not icon:
            return escape_text(match.group(0))
        protected.append(icon)
        return f"\uE000{len(protected) - 1}\uE001"

    def replace_label(match: re.Match[str]) -> str:
        label = match.group(1)
        if iconize_words and label.lower() == "accessorio":
            size = STYLE["cell"]["font_size"] + 4.6
            icon = icon_token(word_icon_files["accessorio"], size)
            if icon:
                protected.append(f"{icon}:")
                return f"\uE000{len(protected) - 1}\uE001"
        color = color_for_semantic_label(label)
        text = escape_text(label.title())
        font = bold_font or font_name("bold")
        token = (
            f'<font name="{font}" color="{color}">'
            f"{text}:</font>"
        )
        protected.append(token)
        return f"\uE000{len(protected) - 1}\uE001"

    working_value = value
    working_value = symbol_pattern.sub(replace_symbol, working_value)
    working_value = inline_icon_pattern.sub(replace_inline_icon, working_value)
    if label_pattern:
        working_value = label_pattern.sub(replace_label, working_value)
    if iconize_words:
        working_value = word_icon_pattern.sub(replace_word_icon, working_value)

    rules = []
    for rule in semantic["rules"]:
        for term in rule["terms"]:
            rules.append((term, rule["color"]))
    rules.sort(key=lambda item: len(item[0]), reverse=True)

    pattern = re.compile(
        "|".join(rf"(?<!\w){re.escape(term)}(?!\w)" for term, _color in rules),
        flags=re.IGNORECASE,
    )
    color_lookup = {term.lower(): color for term, color in rules}

    result: list[str] = []
    last = 0
    for match in pattern.finditer(working_value):
        result.append(escape_text(working_value[last : match.start()]))
        matched = match.group(0)
        color = color_lookup.get(matched.lower(), STYLE["cell"]["text_color"])
        text = escape_text(matched)
        if semantic["bold"] and bold_font:
            result.append(f'<font name="{bold_font}" color="{color}">{text}</font>')
        else:
            result.append(f'<font color="{color}">{text}</font>')
        last = match.end()
    result.append(escape_text(working_value[last:]))
    highlighted = "".join(result)
    for index, replacement in enumerate(protected):
        highlighted = highlighted.replace(escape_text(f"\uE000{index}\uE001"), replacement)
    return highlighted


def color_for_semantic_label(label: str) -> str:
    normalized = label.lower()
    for rule in STYLE["semantic_highlights"]["rules"]:
        if normalized in [term.lower() for term in rule["terms"]] or normalized == rule["name"].lower():
            return rule["color"]
    fallback = {
        "attacco": "#A8322D",
        "movimento": "#2F6F73",
        "difesa": "#3F6F3A",
        "estrazione": "#6B4D9A",
        "accessorio": "#7B5531",
        "accendi": "#9A6B1F",
        "spegni": "#9A6B1F",
        "termina": "#6B4D9A",
        "reazione": "#A8322D",
    }
    return fallback.get(normalized, STYLE["cell"]["text_color"])


def watermark_icon_for_value(value: str) -> Path | None:
    if re.search(r"(?<!\w)L3(?!\w)", value, flags=re.IGNORECASE):
        return ICONS_DIR / "new_symbols" / "11_hex_4_cluster.png"
    if re.search(r"(?<!\w)(?:ES|E5)(?!\w)", value, flags=re.IGNORECASE):
        return ICONS_DIR / "new_symbols" / "10_hex_1_2_3_cluster.png"
    if re.search(r"(?<!\w)(?:raggio|raggi)(?!\w)", value, flags=re.IGNORECASE):
        if re.search(r"^\s*TORCIA\b", value, flags=re.IGNORECASE) or re.search(
            r"\baccendi\b.*\billumina\b", value, flags=re.IGNORECASE
        ):
            return None
        return ICONS_DIR / "new_symbols" / "09_hex_2_2_2.png"
    return None


def add_watermark_if_needed(value: str, paragraph: Any) -> Any:
    icon_path = watermark_icon_for_value(value)
    if icon_path and icon_path.exists():
        return WatermarkedParagraph(
            paragraph,
            icon_path,
            size=STYLE["cell"]["font_size"] * 6.8,
            opacity=0.52,
        )
    return paragraph


def build_sectioned_text(
    value: str,
    paragraph_style: ParagraphStyle,
    changed_label_gap: float = 3.2,
    repeated_label_gap: float = 0,
) -> Table | None:
    lines = [line.strip() for line in value.split("\n") if line.strip()]
    parsed = [split_section_line(line) for line in lines]
    if not any(parsed):
        return None

    rows: list[list[Any]] = []
    style_commands: list[tuple] = [
        ("VALIGN", (0, 0), (0, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (0, -1), 0),
        ("RIGHTPADDING", (0, 0), (0, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]
    previous_label: str | None = None
    label_font = font_name("bold")
    label_size = min(paragraph_style.fontSize, 7.1)

    for row_index, (line, section) in enumerate(zip(lines, parsed, strict=True)):
        if section:
            label, body = section
            label_text = f'<font name="{label_font}" size="{label_size}">{escape_text(label)}:</font>'
            body_text = f"{label_text}&nbsp;&nbsp;{semantic_highlight(body, font_name('bold'))}"
            hanging_indent = section_hanging_indent(label, label_font, label_size)
            if previous_label:
                row_gap = changed_label_gap if label != previous_label else repeated_label_gap
                if row_gap:
                    style_commands.append(("TOPPADDING", (0, row_index), (-1, row_index), row_gap))
            previous_label = label
        else:
            body_text = semantic_highlight(line, font_name("bold"))
            hanging_indent = 0

        row_style = ParagraphStyle(
            f"{paragraph_style.name}Section{row_index}",
            parent=paragraph_style,
            leftIndent=hanging_indent,
            firstLineIndent=-hanging_indent,
        )
        rows.append([Paragraph(body_text, row_style)])

    return Table(
        rows,
        colWidths=[None],
        style=TableStyle(style_commands),
    )


def render_multiline_text(value: str, paragraph_style: ParagraphStyle) -> Any:
    if not value:
        return Paragraph(STYLE["text"]["empty_cell"], paragraph_style)
    sectioned = build_sectioned_text(value, paragraph_style)
    if sectioned:
        return add_watermark_if_needed(value, sectioned)
    paragraph = Paragraph(semantic_highlight(format_section_breaks(value), font_name("bold")), paragraph_style)
    return add_watermark_if_needed(value, paragraph)


def render_creation(value: str, paragraph_styles: dict[str, ParagraphStyle]) -> Any:
    if not value:
        return render_multiline_text(value, paragraph_styles["creation"])

    lines = [line.strip() for line in value.split("\n") if line.strip()]
    name = escape_text(lines[0].title())
    detail_source = format_section_breaks("\n".join(lines[1:]))
    detail_lines = [semantic_highlight(line, font_name("bold")) for line in detail_source.split("\n")]
    creation = STYLE["creation"]
    regular = font_name("regular")
    bold = font_name("hand")
    text = (
        f'<font name="{bold}" size="{creation["name_font_size"]}">'
        f"{name}</font>"
    )
    title_style = ParagraphStyle(
        "MenuCreationTitle",
        parent=paragraph_styles["creation"],
        fontName=bold,
        fontSize=creation["name_font_size"],
        leading=creation["name_leading"],
    )
    title = Paragraph(text, title_style)

    detail_table = None
    if detail_lines:
        detail_style = ParagraphStyle(
            "MenuCreationDetail",
            parent=paragraph_styles["creation"],
            fontName=regular,
            fontSize=creation["detail_font_size"],
            leading=creation["leading"],
        )
        detail_table = build_sectioned_text(
            detail_source,
            detail_style,
            changed_label_gap=2.4,
            repeated_label_gap=1.15,
        )

    if detail_table:
        content: Any = Table(
            [[title], [detail_table]],
            colWidths=[None],
            style=TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (0, 0), 3.4),
                ]
            ),
        )
    else:
        detail = ""
        if detail_lines:
            detail = "<br/>" + "<br/>".join(
                "<br/>" if not line else f'<font name="{regular}" size="{creation["detail_font_size"]}">{line}</font>'
                for line in detail_lines
            )
        content = Paragraph(f"{text}{detail}", paragraph_styles["creation"])

    return add_watermark_if_needed(value, content)


def parse_resources(value: str) -> list[tuple[str, str]]:
    resources: list[tuple[str, str]] = []
    for part in value.split(";"):
        match = re.match(r"^\s*([^:]+?)\s*:\s*(.+?)\s*$", part)
        if match:
            resources.append((match.group(1).strip().lower(), match.group(2).strip()))
    return resources


def render_resources(value: str, paragraph_styles: dict[str, ParagraphStyle]) -> Any:
    resources = parse_resources(value)
    if not resources:
        return render_multiline_text(value, paragraph_styles["materials"])

    def format_resource_amount(name: str, amount: str) -> str:
        if re.fullmatch(r"\d+(?:[.,]\d+)?", amount.strip()):
            return f"{name} {STYLE['resources']['quantity_marker']}{amount}"
        return f"{name} {amount}".strip()

    if not any((ICONS_DIR / f"{name}.png").exists() for name, _amount in resources):
        separator = STYLE["resources"]["separator"]
        text = separator.join(
            semantic_highlight(format_resource_amount(name, amount), font_name("bold"))
            for name, amount in resources
        )
        return Paragraph(text, paragraph_styles["materials"])

    fragments: list[Any] = []
    icon_size = STYLE["icons"]["size"]
    gap = STYLE["icons"]["gap"]
    marker = STYLE["resources"]["quantity_marker"]

    for index, (name, amount) in enumerate(resources):
        icon_path = ICONS_DIR / f"{name}.png"
        if icon_path.exists():
            fragments.append(Image(str(icon_path), width=icon_size, height=icon_size))
            amount_text = f"{marker}{amount}" if re.fullmatch(r"\d+(?:[.,]\d+)?", amount.strip()) else amount
            fragments.append(Paragraph(escape_text(amount_text), paragraph_styles["quantity"]))
        else:
            text = format_resource_amount(name, amount)
            fragments.append(Paragraph(escape_text(text), paragraph_styles["cell"]))
        if index < len(resources) - 1:
            fragments.append(Paragraph("&nbsp;", paragraph_styles["cell"]))

    return Table(
        [fragments],
        colWidths=[None] * len(fragments),
        style=[
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), gap / 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), gap / 2),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ],
    )


def render_cell(column: str, value: str, paragraph_styles: dict[str, ParagraphStyle]) -> Any:
    if column == "MATERIALI":
        return render_resources(value, paragraph_styles)
    if column == "CREAZIONE":
        return render_creation(value, paragraph_styles)
    if column.startswith("U_"):
        return render_multiline_text(value, paragraph_styles["quantity"])
    if column == "POTENZIAMENTO_1":
        return render_multiline_text(value, paragraph_styles["upgrade_1"])
    if column == "POTENZIAMENTO_2":
        return render_multiline_text(value, paragraph_styles["upgrade_2"])
    if column == "POTENZIAMENTO_3":
        return render_multiline_text(value, paragraph_styles["upgrade_3"])
    return render_multiline_text(value, paragraph_styles["cell"])


def classify_legend_symbol(label: str) -> str:
    normalized = label.strip().upper()
    mapping = {
        "◎": "use",
        "@": "use",
        "U": "use",
        "⚔": "attack",
        "ATTACCO": "attack",
        "↪": "movement",
        "MOVIMENTO": "movement",
        "O": "draw",
        "▽": "defense",
        "✱": "special",
        "*": "special",
        "G": "range",
        "ACCESSORIO": "accessory",
    }
    return mapping.get(normalized, "generic")


def get_legend_icon(symbol_type: str, label: str, paragraph_styles: dict[str, ParagraphStyle]) -> Any:
    icon_file = STYLE["legend_icons"].get(symbol_type)
    if icon_file:
        icon_path = ICONS_DIR / icon_file
        if icon_path.exists():
            size = STYLE["legend_icons"]["size"]
            return Image(str(icon_path), width=size, height=size)
    if symbol_type == "defense":
        size = STYLE["legend_icons"]["size"]
        drawing = Drawing(size, size)
        drawing.add(
            Polygon(
                [1, size - 2, size / 2, 1, size - 1, size - 2],
                strokeColor=hex_color(STYLE["footer"]["text_color"]),
                fillColor=None,
                strokeWidth=1.1,
            )
        )
        return drawing
    return Paragraph(escape_text(label), paragraph_styles["footer_symbol"])


def parse_legend_entries(legend_text: str) -> list[LegendEntry]:
    body = re.sub(r"^\s*LEGENDA\s*:?", "", legend_text, flags=re.IGNORECASE).strip()
    compact = re.sub(r"\s*\|\s*", "  ", body.replace("\n", " "))
    compact = re.sub(r"(?<!\w)accessorio\s*:", "ACCESSORIO:", compact, flags=re.IGNORECASE)
    token_pattern = r"(◎|⚔|↪|▽|✱|@|\*|(?<!\w)U(?=\s*:)|(?<!\w)(?:ATTACCO|MOVIMENTO|ACCESSORIO|G|O)(?!\w))"
    matches = list(re.finditer(token_pattern, compact))
    entries: list[LegendEntry] = []

    for index, match in enumerate(matches):
        label = match.group(1).upper() if match.group(1).isalpha() else match.group(1)
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(compact)
        description = re.sub(r"\s+", " ", compact[start:end]).strip(" =:-\t")
        if description:
            entries.append(
                LegendEntry(
                    symbol_type=classify_legend_symbol(label),
                    label=label,
                    description=prettify_line(description),
                )
            )

    return entries


def render_legend_item(entry: LegendEntry, paragraph_styles: dict[str, ParagraphStyle]) -> Table:
    footer = STYLE["footer"]
    symbol_width = STYLE["legend_icons"]["symbol_width"]
    icon = get_legend_icon(entry.symbol_type, entry.label, paragraph_styles)
    description = Paragraph(
        semantic_highlight(entry.description, font_name("bold"), iconize_words=False),
        paragraph_styles["footer"],
    )
    table = Table(
        [[icon, description]],
        colWidths=[symbol_width, None],
        style=[
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 0),
            ("RIGHTPADDING", (0, 0), (-1, -1), footer["item_gap"] / 2),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ],
    )
    return table


def build_footer_legend(legend_text: str, paragraph_styles: dict[str, ParagraphStyle], width: float) -> Table:
    footer = STYLE["footer"]
    entries = parse_legend_entries(legend_text)

    title = Paragraph("LEGENDA", paragraph_styles["footer_title"])
    if not entries:
        body: list[list[Any]] = [[Paragraph(escape_text(legend_text), paragraph_styles["footer"])]]
        col_widths = [width]
    else:
        columns = footer["columns"]
        item_width = width / columns
        rendered = [render_legend_item(entry, paragraph_styles) for entry in entries]
        body = []
        for row_start in range(0, len(rendered), columns):
            row = rendered[row_start : row_start + columns]
            row.extend([""] * (columns - len(row)))
            body.append(row)
        col_widths = [item_width] * columns

    data = [[title]] + body
    if entries:
        data[0].extend([""] * (footer["columns"] - 1))
    legend_table = Table(
        data,
        colWidths=[width] if not entries else col_widths,
        style=[
            ("SPAN", (0, 0), (-1, 0)),
            ("BACKGROUND", (0, 0), (-1, -1), hex_color(footer["background"])),
            ("BOX", (0, 0), (-1, -1), footer["border_width"], hex_color(footer["border_color"])),
            ("LINEABOVE", (0, 1), (-1, 1), 0.35, hex_color(footer["border_color"])),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), footer["padding"]),
            ("RIGHTPADDING", (0, 0), (-1, -1), footer["padding"]),
            ("TOPPADDING", (0, 0), (-1, -1), footer["padding"] / 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), footer["padding"] / 2),
        ],
    )
    return legend_table


def build_table_data(recipes: list[RecipeRow], paragraph_styles: dict[str, ParagraphStyle]) -> list[list[Any]]:
    data: list[list[Any]] = [
        [Paragraph(escape_text(header.upper()), paragraph_styles["header"]) for header in DISPLAY_HEADERS]
    ]

    for recipe in recipes:
        data.append(
            [
                render_cell(column, recipe.values[column], paragraph_styles)
                for column in REQUIRED_COLUMNS
            ]
        )
    return data


def build_column_widths(available_width: float) -> list[float]:
    relative = STYLE["columns"]["relative_widths"]
    total = sum(relative)
    return [available_width * item / total for item in relative]


def build_table_style(row_count: int) -> TableStyle:
    colors_style = STYLE["colors"]
    cell = STYLE["cell"]
    table = STYLE["table"]
    header = STYLE["header"]

    commands: list[tuple] = [
        ("BACKGROUND", (0, 0), (-1, 0), hex_color(header["background"])),
        ("TEXTCOLOR", (0, 0), (-1, 0), hex_color(header["text_color"])),
        ("ALIGN", (0, 0), (-1, 0), header["align"]),
        ("VALIGN", (0, 0), (-1, -1), cell["valign"]),
        ("FONTNAME", (0, 0), (-1, 0), font_name("bold")),
        ("GRID", (0, 0), (-1, -1), table["inner_grid_width"], hex_color(table["border_color"])),
        ("BOX", (0, 0), (-1, -1), table["outer_grid_width"], hex_color(table["border_color"])),
        ("LINEBELOW", (0, 0), (-1, 0), 0.65, hex_color(table["border_color"])),
        ("LINEBELOW", (0, 0), (-1, 0), 0.25, hex_color(table["border_color"])),
        ("LEFTPADDING", (0, 0), (-1, -1), cell["padding_left"]),
        ("RIGHTPADDING", (0, 0), (-1, -1), cell["padding_right"]),
        ("TOPPADDING", (0, 0), (-1, -1), cell["padding_top"]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), cell["padding_bottom"]),
    ]

    column_backgrounds = [
        (0, STYLE["creation"]["background"]),
        (1, STYLE["quantity"]["background"]),
        (2, STYLE["materials"]["background"]),
        (3, STYLE["quantity"]["background"]),
        (4, STYLE["upgrade_1"]["background"]),
        (5, STYLE["quantity"]["background"]),
        (6, STYLE["upgrade_2"]["background"]),
        (7, STYLE["quantity"]["background"]),
        (8, STYLE["upgrade_3"]["background"]),
        (9, STYLE["quantity"]["background"]),
    ]
    for column, background in column_backgrounds:
        if background:
            commands.append(("BACKGROUND", (column, 1), (column, -1), hex_color(background)))

    for row in range(1, row_count):
        if row % 2 == 0:
            if STYLE["creation"]["background"]:
                commands.append(("BACKGROUND", (0, row), (0, row), hex_color(colors_style["creation_alt"])))
            if STYLE["quantity"]["background"]:
                commands.append(("BACKGROUND", (1, row), (1, row), hex_color(colors_style["quantity_alt"])))
                commands.append(("BACKGROUND", (3, row), (3, row), hex_color(colors_style["quantity_alt"])))
                commands.append(("BACKGROUND", (5, row), (5, row), hex_color(colors_style["quantity_alt"])))
                commands.append(("BACKGROUND", (7, row), (7, row), hex_color(colors_style["quantity_alt"])))
                commands.append(("BACKGROUND", (9, row), (9, row), hex_color(colors_style["quantity_alt"])))

    return TableStyle(commands)


def build_table(
    table_data: list[list[Any]],
    col_widths: list[float],
    available_width: float,
    available_height: float,
) -> Table:
    repeat_rows = 1 if STYLE["table"]["repeat_header"] else 0
    table = RecipeTable(
        table_data,
        colWidths=col_widths,
        rowHeights=None,
        repeatRows=repeat_rows,
        splitByRow=True,
    )
    table_style = build_table_style(len(table_data))
    table.setStyle(table_style)
    _width, natural_height = table.wrap(available_width, available_height)

    target_height = available_height * STYLE["table"]["fill_height_ratio"]
    if natural_height < target_height and len(table_data) > 1:
        natural_row_heights = list(table._rowHeights)
        extra = target_height - natural_height
        extra_per_data_row = extra / (len(table_data) - 1)
        row_heights = [natural_row_heights[0]] + [
            height + extra_per_data_row for height in natural_row_heights[1:]
        ]
        table = RecipeTable(
            table_data,
            colWidths=col_widths,
            rowHeights=row_heights,
            repeatRows=repeat_rows,
            splitByRow=True,
        )
        table.setStyle(table_style)

    return table


def draw_background(canvas: Any, doc: SimpleDocTemplate) -> None:
    page_color = STYLE["page"]["background_color"]
    canvas.saveState()
    canvas.setFillColor(hex_color(page_color))
    canvas.rect(0, 0, doc.pagesize[0], doc.pagesize[1], fill=1, stroke=0)

    background_path = PAPER_TEXTURE if PAPER_TEXTURE.exists() else GENERATED_PAPER_TEXTURE
    if background_path.exists():
        try:
            canvas.drawImage(
                str(background_path),
                0,
                0,
                width=doc.pagesize[0],
                height=doc.pagesize[1],
                preserveAspectRatio=False,
                mask="auto",
            )
        except Exception:
            pass

    canvas.restoreState()


def create_preview(pdf_file: Path, preview_file: Path = PREVIEW_FILE) -> Path | None:
    pdftoppm = shutil.which("pdftoppm")
    if not pdftoppm:
        return None

    preview_file.parent.mkdir(parents=True, exist_ok=True)
    output_stem = preview_file.with_suffix("")
    result = subprocess.run(
        [pdftoppm, "-png", "-singlefile", "-r", "150", str(pdf_file), str(output_stem)],
        check=False,
        capture_output=True,
        text=True,
    )
    if result.returncode != 0:
        return None
    return preview_file


def build_pdf(recipes: list[RecipeRow], legend_text: str, output_file: Path = OUTPUT_FILE) -> Path:
    output_file.parent.mkdir(parents=True, exist_ok=True)
    register_fonts()
    generate_paper_artwork(GENERATED_PAPER_TEXTURE, STYLE, DOWNLOADED_PAPER_TEXTURE, BOTANICAL_WATERCOLOR)
    ensure_legend_icons(ICONS_DIR, STYLE)

    page_size = landscape(A4)
    margin = STYLE["page"]["margin_mm"] * mm
    doc = SimpleDocTemplate(
        str(output_file),
        pagesize=page_size,
        leftMargin=margin,
        rightMargin=margin,
        topMargin=margin,
        bottomMargin=margin,
        title="Menu Ricette",
        author="menu_gioco",
    )

    paragraph_styles = create_paragraph_styles()
    footer = build_footer_legend(legend_text, paragraph_styles, doc.width)
    _footer_width, footer_height = footer.wrap(doc.width, doc.height)
    footer_spacing = STYLE["footer"]["spacing"]
    table_available_height = max(doc.height - footer_height - footer_spacing, doc.height * 0.6)

    table_data = build_table_data(recipes, paragraph_styles)
    col_widths = build_column_widths(doc.width)
    table = build_table(table_data, col_widths, doc.width, table_available_height)

    try:
        doc.build([table, Spacer(1, footer_spacing), footer], onFirstPage=draw_background, onLaterPages=draw_background)
    except PermissionError as exc:
        raise MenuGenerationError(f"File PDF non scrivibile: {output_file}") from exc
    except OSError as exc:
        raise MenuGenerationError(f"Impossibile creare il PDF: {output_file}") from exc

    return output_file


def main() -> int:
    try:
        print(f"[1/4] Lettura Excel: {INPUT_FILE.name}...")
        recipes = load_workbook_data()
        legend_text = load_footer_legend()
        print(f"[2/4] {len(recipes)} ricette caricate + legenda")
        print("[3/4] Generazione layout A4...")
        output_file = build_pdf(recipes, legend_text)
        preview_file = create_preview(output_file)
        print("[4/4] PDF creato correttamente")
        print()
        print(output_file.relative_to(Path(__file__).resolve().parent))
        if preview_file:
            print(preview_file.relative_to(Path(__file__).resolve().parent))
        if INPUT_FILE != DEFAULT_INPUT_FILE:
            print(f"Input usato: {INPUT_FILE.name}")
        return 0
    except MenuGenerationError as exc:
        print(f"Errore: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
